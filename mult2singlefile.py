# import shutil
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter


class setting_rule:
    def __init__(self, sour_colnum, dst_coordinate):
        self.sour_colnum = sour_colnum
        self.dst_coordinate = dst_coordinate


# CONF_ADDR = "C:\\Users\\h\\Downloads\\config.xlsx"
SOURCE_FILE = "C:\\Users\\h\\Downloads\\source.xlsx"
TEMPLATE_FILE = "C:\\Users\\h\\Downloads\\template.xlsx"
OUT_ADDR = "C:\\Users\\h\\Downloads\\output\\"


# 配置规则：源文件中的位置，目标文件的位置，备注
setting_list = [
    ['C', 'B3', '姓名'],
    ['D', 'D3', '身份证号'],
    ['E', 'F3', '行政职级待遇'],
    ['F', 'C4', '是否分配安置住房'],
    ['G', 'B6', '货币补差情况'],
    ['H', 'B9', '现住房类型'],
    ['I', 'C13', '小区名称'],
    ['J', 'C14', '栋号'],
    ['K', 'C15', '单元号'],
    ['L', 'C16', '室号'],
    ['M', 'D17', '建筑面积'],
    ['N', 'F17', '入住时间'],
    ['O', 'C18', '是否办理购房手续'],
    ['P', 'F18', '购房时间'],
    ['Q', 'C19', '产权证办理情况'],
    ['R', 'C20', '现居住人类别'],
    ['S', 'C21', '现居住人姓名'],
    ['T', 'C22', '过户情况'],
    ['U', 'C23', '使用情况'],
    ['V', 'C24', '分散安置住房地址'],
    ['W', 'C25', '现住房地址'],
    ['X', 'F25', '建筑面积'],
    ['AM', 'B26', '备注']
]

# 生成易读规则
rules = [setting_rule(single_setting[0], single_setting[1])
         for single_setting in setting_list]


# 行范围，正式： 9, 700
sta_line_num, end_lin_num = 9, 700


# 源列表
s_wb = load_workbook(SOURCE_FILE, read_only=True)
sour_sheet = s_wb[s_wb.sheetnames[0]]


# 按行处理
for i in range(sta_line_num, end_lin_num+1):
    col = str(i)

    d_wb = load_workbook(TEMPLATE_FILE)
    # 目标表单
    dst_sheet = d_wb[d_wb.sheetnames[0]]

    # 按照规则填空
    for r in rules:
        dst_sheet[r.dst_coordinate] = sour_sheet[r.sour_colnum+col].value

    # 最终文件名
    dst_end_name = OUT_ADDR+sour_sheet['D'+col].value+'.xlsx'

    # 保存到身份证号文件名
    d_wb.save(dst_end_name)
